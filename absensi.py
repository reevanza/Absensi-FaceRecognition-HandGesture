import cv2
import dlib
import math
import mediapipe as mp
import os
import numpy as np
import random
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from collections import Counter, deque

# -----------------------
# Fungsi deteksi wajah (detail)
# -----------------------
def euclidean_dist(p1, p2):
    return math.dist(p1, p2)

def mouth_aspect_ratio(landmarks):
    p48 = (landmarks.part(48).x, landmarks.part(48).y)
    p54 = (landmarks.part(54).x, landmarks.part(54).y)
    p51 = (landmarks.part(51).x, landmarks.part(51).y)
    p57 = (landmarks.part(57).x, landmarks.part(57).y)
    p49 = (landmarks.part(49).x, landmarks.part(49).y)
    p55 = (landmarks.part(55).x, landmarks.part(55).y)
    p53 = (landmarks.part(53).x, landmarks.part(53).y)
    p59 = (landmarks.part(59).x, landmarks.part(59).y)
    mar = (euclidean_dist(p51, p57) + euclidean_dist(p53, p59) + euclidean_dist(p49, p55)) / (3 * euclidean_dist(p48, p54))
    return mar

def smile_aspect_ratio(landmarks):
    p48 = (landmarks.part(48).x, landmarks.part(48).y)
    p54 = (landmarks.part(54).x, landmarks.part(54).y)
    p51 = (landmarks.part(51).x, landmarks.part(51).y)
    p57 = (landmarks.part(57).x, landmarks.part(57).y)
    sar = euclidean_dist(p48, p54) / euclidean_dist(p51, p57)
    return sar

MOUTH_AR_THRESH = 0.55
SMILE_AR_THRESH = 1.7

# Load predictor dlib
PREDICTOR_PATH = "shape_predictor_68_face_landmarks.dat"
predictor = None
detector_dlib = None
if not os.path.exists(PREDICTOR_PATH):
    print(f"[WARNING] {PREDICTOR_PATH} tidak ditemukan.")
else:
    try:
        detector_dlib = dlib.get_frontal_face_detector()
        predictor = dlib.shape_predictor(PREDICTOR_PATH)
    except:
        pass

# -----------------------
# Gesture detection
# -----------------------
mp_hands = mp.solutions.hands
hands = mp_hands.Hands()
mp_draw = mp.solutions.drawing_utils

def recognize_gesture(hand_landmarks):
    ujung_jempol = hand_landmarks.landmark[mp_hands.HandLandmark.THUMB_TIP]
    ujung_telunjuk = hand_landmarks.landmark[mp_hands.HandLandmark.INDEX_FINGER_TIP]
    ujung_tengah = hand_landmarks.landmark[mp_hands.HandLandmark.MIDDLE_FINGER_TIP]
    ujung_manis = hand_landmarks.landmark[mp_hands.HandLandmark.RING_FINGER_TIP]
    ujung_kelingking = hand_landmarks.landmark[mp_hands.HandLandmark.PINKY_TIP]
    
    pangkal_jempol = hand_landmarks.landmark[mp_hands.HandLandmark.THUMB_MCP]
    pangkal_telunjuk = hand_landmarks.landmark[mp_hands.HandLandmark.INDEX_FINGER_MCP]
    pangkal_tengah = hand_landmarks.landmark[mp_hands.HandLandmark.MIDDLE_FINGER_MCP]
    pangkal_manis = hand_landmarks.landmark[mp_hands.HandLandmark.RING_FINGER_MCP]
    pangkal_kelingking = hand_landmarks.landmark[mp_hands.HandLandmark.PINKY_MCP]
    
    if (ujung_jempol.y < ujung_telunjuk.y and
        ujung_jempol.y < ujung_tengah.y and
        ujung_jempol.y < ujung_manis.y and
        ujung_jempol.y < ujung_kelingking.y):
        return "Thumbs Up"
    elif (ujung_telunjuk.y < pangkal_telunjuk.y and
          ujung_tengah.y < pangkal_tengah.y and
          ujung_manis.y > pangkal_manis.y and
          ujung_kelingking.y > pangkal_kelingking.y):
        return "Peace"
    elif (ujung_telunjuk.y < pangkal_telunjuk.y and
          ujung_kelingking.y < pangkal_kelingking.y and
          ujung_tengah.y > pangkal_tengah.y and
          ujung_manis.y > pangkal_manis.y):
        return "Rock"
    elif (abs(ujung_jempol.x - ujung_telunjuk.x) < 0.08 and # Dilonggarkan dari 0.05
            abs(ujung_jempol.y - ujung_telunjuk.y) < 0.08 and # Dilonggarkan dari 0.05
            ujung_tengah.y < pangkal_tengah.y and
            ujung_manis.y < pangkal_manis.y and
            ujung_kelingking.y < pangkal_kelingking.y):
        return "OK"
    elif (ujung_jempol.y < pangkal_jempol.y and
          ujung_telunjuk.y < pangkal_telunjuk.y and
          ujung_tengah.y < pangkal_tengah.y and
          ujung_manis.y < pangkal_manis.y and
          ujung_kelingking.y < pangkal_kelingking.y):
        return "Hi"
    return "Gesture tidak diketahui"

def detect_hand_gesture(image):
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    results = hands.process(image_rgb)
    gesture_found = None
    if results.multi_hand_landmarks:
        for hand_landmarks in results.multi_hand_landmarks:
            gesture_found = recognize_gesture(hand_landmarks)
            mp_draw.draw_landmarks(image, hand_landmarks, mp_hands.HAND_CONNECTIONS)
    return image, gesture_found

# -----------------------
# Face recognizer dengan preprocessing minimal
# -----------------------
DATASET_DIR = "dataset"
IMG_SIZE = (200, 200)

def preprocess_face(face_img):
    """Preprocessing MINIMAL - hanya normalisasi pencahayaan"""
    face_img = cv2.equalizeHist(face_img)
    return face_img

def train_recognizer(dataset_dir=DATASET_DIR):
    if not os.path.exists(dataset_dir):
        return None, {}
    
    faces, labels, label_ids = [], [], {}
    current_id = 0
    
    print("[INFO] Training face recognizer...")
    for person_name in sorted(os.listdir(dataset_dir)):
        path = os.path.join(dataset_dir, person_name)
        if not os.path.isdir(path): 
            continue
            
        label_ids[person_name] = current_id
        img_count = 0
        
        for img_name in os.listdir(path):
            img_path = os.path.join(path, img_name)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            
            if img is not None:
                img_resized = cv2.resize(img, IMG_SIZE)
                img_processed = preprocess_face(img_resized)
                
                faces.append(img_processed)
                labels.append(current_id)
                img_count += 1
        
        print(f"  - {person_name}: {img_count} gambar")
        current_id += 1
    
    if not faces: 
        return None, {}
    
    # LBPH dengan parameter default (paling stabil)
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.train(faces, np.array(labels))
    
    print(f"[INFO] Training selesai. Total orang: {current_id}")
    return recognizer, {v:k for k,v in label_ids.items()}

# -----------------------
# Excel helper
# -----------------------
EXCEL_PATH = "absensi.xlsx"
def append_to_excel(name, waktu, mulut, gesture, confidence):
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nama", "Waktu", "Mulut", "Gesture", "Confidence"])
    ws.append([name, waktu, mulut, gesture, f"{confidence:.1f}"])
    wb.save(EXCEL_PATH)

# -----------------------
# Random combination
# -----------------------
mouths = ["Senyum", "Mulut terbuka"]
gestures_list = ["Thumbs Up", "Peace", "Rock", "OK", "Hi"]

chosen_mouth = random.choice(mouths)
chosen_gesture = random.choice(gestures_list)
print(f"[INFO] Kombinasi acak: Mulut={chosen_mouth} | Gesture={chosen_gesture}")

# -----------------------
# Initialize
# -----------------------
recognizer, id_to_name = train_recognizer()
haar = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
TEXT_COLOR = (0, 0, 0)
last_absen_time = {}
COOLDOWN = 5  # minutes

# CONFIDENCE SETTINGS - CRITICAL!
# Semakin RENDAH = semakin YAKIN (0 = perfect match)
CONFIDENCE_THRESHOLD = 70
VOTING_FRAMES = 20         
VOTING_THRESHOLD = 0.9   

# Queue untuk voting
face_predictions = deque(maxlen=VOTING_FRAMES)

cap = cv2.VideoCapture(0)
cap.set(3, 1280)
cap.set(4, 720)

print(f"\n[INFO] Confidence threshold: {CONFIDENCE_THRESHOLD}")
print(f"[INFO] Voting: {int(VOTING_THRESHOLD*100)}% dari {VOTING_FRAMES} frames")
print("[INFO] Tekan 'q' untuk keluar\n")

# -----------------------
# Main loop
# -----------------------
while True:
    ret, frame = cap.read()
    if not ret: 
        break
        
    frame_display = frame.copy()
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    frame_display, gesture_found = detect_hand_gesture(frame_display)

    mouth_detected = None
    faces_for_recog = []

    if detector_dlib and predictor:
        faces = detector_dlib(gray)
        for face in faces:
            landmarks = predictor(gray, face)
            mar = mouth_aspect_ratio(landmarks)
            sar = smile_aspect_ratio(landmarks)
            if sar > SMILE_AR_THRESH:
                mouth_detected = "Senyum"
            elif mar > MOUTH_AR_THRESH:
                mouth_detected = "Mulut terbuka"
            x1, y1, x2, y2 = face.left(), face.top(), face.right(), face.bottom()
            faces_for_recog.append((max(0,x1), max(0,y1), x2-x1, y2-y1))
            
            for n in range(0, 68):
                x = landmarks.part(n).x
                y = landmarks.part(n).y
                cv2.circle(frame_display, (x, y), 1, (0, 255, 255), -1)
    else:
        faces = haar.detectMultiScale(gray, 1.3, 5)
        for (x,y,w,h) in faces:
            faces_for_recog.append((x,y,w,h))

    # --- Text UI ---
    y_text = 40
    cv2.putText(frame_display, "INSTRUKSI:", (30, y_text), cv2.FONT_HERSHEY_SIMPLEX, 0.95, TEXT_COLOR, 2)
    y_text += 40
    cv2.putText(frame_display, f"- Mulut: {chosen_mouth}", (30, y_text), cv2.FONT_HERSHEY_SIMPLEX, 0.9, TEXT_COLOR, 2)
    y_text += 40
    cv2.putText(frame_display, f"- Gesture: {chosen_gesture}", (30, y_text), cv2.FONT_HERSHEY_SIMPLEX, 0.9, TEXT_COLOR, 2)
    y_text += 50
    cv2.putText(frame_display, f"Detected - Mulut: {mouth_detected or '-'} | Gesture: {gesture_found or '-'}", 
                (30, y_text), cv2.FONT_HERSHEY_SIMPLEX, 0.85, TEXT_COLOR, 2)

    # --- Check match ---
    mouth_ok = (mouth_detected == chosen_mouth)
    gesture_ok = (gesture_found == chosen_gesture)

    # Proses setiap wajah
    current_frame_prediction = None
    
    for (x,y,w,h) in faces_for_recog:
        face_roi = gray[y:y+h, x:x+w]
        face_rs = cv2.resize(face_roi, IMG_SIZE)
        face_processed = preprocess_face(face_rs)
        
        identified_name = "Unknown"
        conf = None
        box_color = (0, 0, 255)
        
        if recognizer:
            try:
                label, conf = recognizer.predict(face_processed)
                
                # Hanya terima jika confidence di bawah threshold
                if conf < CONFIDENCE_THRESHOLD:
                    predicted_name = id_to_name.get(label, "Unknown")
                    current_frame_prediction = predicted_name
                    box_color = (0, 255, 0)
                else:
                    current_frame_prediction = "Unknown"
                    box_color = (0, 0, 255)
                    
            except Exception as e:
                print(f"[ERROR] {e}")
                current_frame_prediction = "Unknown"
        
        # Tambahkan ke voting queue
        if current_frame_prediction:
            face_predictions.append(current_frame_prediction)
        
        # Hitung voting dari 10 frame terakhir
        if len(face_predictions) >= VOTING_FRAMES:
            vote_counts = Counter(face_predictions)
            most_common = vote_counts.most_common(1)[0]
            voted_name, vote_count = most_common
            vote_percentage = vote_count / VOTING_FRAMES
            
            # Hanya terima jika voting >= threshold DAN bukan Unknown
            if vote_percentage >= VOTING_THRESHOLD and voted_name != "Unknown":
                identified_name = voted_name
                box_color = (0, 255, 0)
            else:
                identified_name = "Unknown"
                box_color = (0, 0, 255)
            
            # Tampilkan voting info
            cv2.putText(frame_display, f"Vote: {voted_name} ({vote_count}/{VOTING_FRAMES})", 
                       (30, 180), cv2.FONT_HERSHEY_SIMPLEX, 0.8, TEXT_COLOR, 2)
        else:
            identified_name = "Menunggu..."
            box_color = (128, 128, 128)

        # Gambar kotak dan info
        cv2.rectangle(frame_display, (x,y), (x+w,y+h), box_color, 2)
        
        text = f"{identified_name}"
        if conf:
            text += f" ({conf:.1f})"
        
        text_size = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)[0]
        cv2.rectangle(frame_display, (x, y-30), (x+text_size[0]+10, y), box_color, -1)
        cv2.putText(frame_display, text, (x+5, y-10), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

        # Proses absensi
        if mouth_ok and gesture_ok and identified_name != "Unknown" and identified_name != "Menunggu...":
            now = datetime.now()
            last_time = last_absen_time.get(identified_name)
            
            if not last_time or (now - last_time) >= timedelta(minutes=COOLDOWN):
                append_to_excel(identified_name, now.strftime("%Y-%m-%d %H:%M:%S"), 
                              chosen_mouth, chosen_gesture, conf if conf else 0)
                last_absen_time[identified_name] = now
                
                cv2.putText(frame_display, f"✅ Absensi: {identified_name}", 
                           (30, 300), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 0), 3)
                print(f"[ABSEN] {identified_name} - Confidence: {conf:.1f}")
                
                # Reset voting setelah absen
                face_predictions.clear()
            else:
                sisa = int((timedelta(minutes=COOLDOWN)-(now-last_time)).total_seconds())
                cv2.putText(frame_display, f"⏳ Tunggu {sisa//60}:{sisa%60:02d}", 
                           (30, 300), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 165, 255), 3)

    cv2.imshow("Face+Hand Attendance (Voting System)", frame_display)
    if cv2.waitKey(1) & 0xFF == ord('q'): 
        break

cap.release()
cv2.destroyAllWindows()
print("\n[INFO] Program selesai.")