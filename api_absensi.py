from fastapi import FastAPI, UploadFile, File
from fastapi.responses import JSONResponse
import cv2
import dlib
import mediapipe as mp
import numpy as np
import math, os, random
from datetime import datetime
from openpyxl import Workbook, load_workbook

def euclidean_dist(p1, p2):
    return math.dist(p1, p2)

def mouth_aspect_ratio(landmarks):
    p48 = (landmarks.part(48).x, landmarks.part(48).y)
    p54 = (landmarks.part(54).x, landmarks.part(54).y)
    p51 = (landmarks.part(51).x, landmarks.part(51).y)
    p57 = (landmarks.part(57).x, landmarks.part(57).y)
    p49 = (landmarks.part(49).x, landmarks.part(49).y)
    p55 = (landmarks.part(55).x, landmarks.part(55).y)
    p53 = (landmarks.part(53).x, landmarks.part(53).y)
    p59 = (landmarks.part(59).x, landmarks.part(59).y)
    mar = (euclidean_dist(p51, p57) + euclidean_dist(p53, p59) + euclidean_dist(p49, p55)) / (3 * euclidean_dist(p48, p54))
    return mar

def smile_aspect_ratio(landmarks):
    p48 = (landmarks.part(48).x, landmarks.part(48).y)
    p54 = (landmarks.part(54).x, landmarks.part(54).y)
    p51 = (landmarks.part(51).x, landmarks.part(51).y)
    p57 = (landmarks.part(57).x, landmarks.part(57).y)
    sar = euclidean_dist(p48, p54) / euclidean_dist(p51, p57)
    return sar

MOUTH_AR_THRESH = 0.6
SMILE_AR_THRESH = 1.8

PREDICTOR_PATH = "shape_predictor_68_face_landmarks.dat"
predictor = None
detector_dlib = None
if os.path.exists(PREDICTOR_PATH):
    detector_dlib = dlib.get_frontal_face_detector()
    predictor = dlib.shape_predictor(PREDICTOR_PATH)
else:
    print(f"[WARNING] {PREDICTOR_PATH} tidak ditemukan. Hanya gesture aktif.")

mp_hands = mp.solutions.hands
hands = mp_hands.Hands()
mp_draw = mp.solutions.drawing_utils

def recognize_gesture(hand_landmarks):
    ujung_jempol = hand_landmarks.landmark[mp_hands.HandLandmark.THUMB_TIP]
    ujung_telunjuk = hand_landmarks.landmark[mp_hands.HandLandmark.INDEX_FINGER_TIP]
    ujung_tengah = hand_landmarks.landmark[mp_hands.HandLandmark.MIDDLE_FINGER_TIP]
    ujung_manis = hand_landmarks.landmark[mp_hands.HandLandmark.RING_FINGER_TIP]
    ujung_kelingking = hand_landmarks.landmark[mp_hands.HandLandmark.PINKY_TIP]
    
    pangkal_jempol = hand_landmarks.landmark[mp_hands.HandLandmark.THUMB_MCP]
    pangkal_telunjuk = hand_landmarks.landmark[mp_hands.HandLandmark.INDEX_FINGER_MCP]
    pangkal_tengah = hand_landmarks.landmark[mp_hands.HandLandmark.MIDDLE_FINGER_MCP]
    pangkal_manis = hand_landmarks.landmark[mp_hands.HandLandmark.RING_FINGER_MCP]
    pangkal_kelingking = hand_landmarks.landmark[mp_hands.HandLandmark.PINKY_MCP]
    
    if (ujung_jempol.y < ujung_telunjuk.y and ujung_jempol.y < ujung_tengah.y and ujung_jempol.y < ujung_manis.y and ujung_jempol.y < ujung_kelingking.y):
        return "Thumbs Up"
    elif (ujung_telunjuk.y < pangkal_telunjuk.y and ujung_tengah.y < pangkal_tengah.y and ujung_manis.y > pangkal_manis.y and ujung_kelingking.y > pangkal_kelingking.y):
        return "Peace"
    elif (ujung_telunjuk.y < pangkal_telunjuk.y and ujung_kelingking.y < pangkal_kelingking.y and ujung_tengah.y > pangkal_tengah.y and ujung_manis.y > pangkal_manis.y):
        return "Rock"
    elif (abs(ujung_jempol.x - ujung_telunjuk.x) < 0.05 and abs(ujung_jempol.y - ujung_telunjuk.y) < 0.05):
        return "OK"
    elif (ujung_jempol.y < pangkal_jempol.y and ujung_telunjuk.y < pangkal_telunjuk.y and ujung_tengah.y < pangkal_tengah.y and ujung_manis.y < pangkal_manis.y and ujung_kelingking.y < pangkal_kelingking.y):
        return "Hi"
    return None

def detect_hand_gesture(image):
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    results = hands.process(image_rgb)
    if results.multi_hand_landmarks:
        for hand_landmarks in results.multi_hand_landmarks:
            gesture = recognize_gesture(hand_landmarks)
            return gesture
    return None

EXCEL_PATH = "absensi.xlsx"

def append_to_excel(name, waktu, mulut, gesture):
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nama", "Waktu", "Mulut", "Gesture"])
    ws.append([name, waktu, mulut, gesture])
    wb.save(EXCEL_PATH)

app = FastAPI(title="API Absensi Face & Gesture", version="1.0")

@app.get("/")
def home():
    return {"message": "API Absensi aktif. Gunakan /docs untuk uji API."}

@app.post("/absensi")
async def proses_absensi(file: UploadFile = File(...)):
    try:
        image_bytes = await file.read()
        nparr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame is None:
            return JSONResponse(content={"error": "Gambar tidak valid."}, status_code=400)

        gesture = detect_hand_gesture(frame)

        mouth_status = "-"
        if predictor and detector_dlib:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = detector_dlib(gray)
            for face in faces:
                landmarks = predictor(gray, face)
                mar = mouth_aspect_ratio(landmarks)
                sar = smile_aspect_ratio(landmarks)
                if sar > SMILE_AR_THRESH:
                    mouth_status = "Senyum"
                elif mar > MOUTH_AR_THRESH:
                    mouth_status = "Mulut terbuka"
                break 

        if gesture is None:
            gesture = "-"
        if mouth_status is None or isinstance(mouth_status, float) and (np.isnan(mouth_status) or np.isinf(mouth_status)):
            mouth_status = "-"
    
        waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        append_to_excel("User", waktu, mouth_status, gesture or "-")

        return {
            "status": "sukses",
            "waktu": waktu,
            "mulut": str(mouth_status),
            "gesture": str(gesture)
        }
    
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.get("/data")
def ambil_data():
    import pandas as pd
    if not os.path.exists(EXCEL_PATH):
        return {"data": []}
    df = pd.read_excel(EXCEL_PATH)
    return {"data": df.to_dict(orient="records")}
