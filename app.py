from flask import Flask, render_template, request, jsonify
import cv2, numpy as np, base64, os
from datetime import datetime, timedelta
import dlib, math, random
import mediapipe as mp
from openpyxl import Workbook, load_workbook

# === Konfigurasi dasar ===
app = Flask(__name__)
DATASET_DIR = "dataset"
EXCEL_PATH = "absensi.xlsx"
IMG_SIZE = (200, 200)
COOLDOWN = 5  # menit
TEXT_COLOR = (0, 0, 0)

# -----------------------
# Bagian: Fungsi wajah (dari faceHandAttendance_v4.py)
# -----------------------
def euclidean_dist(p1, p2):
    return math.dist(p1, p2)

def mouth_aspect_ratio(landmarks):
    p48 = (landmarks.part(48).x, landmarks.part(48).y)
    p54 = (landmarks.part(54).x, landmarks.part(54).y)
    p51 = (landmarks.part(51).x, landmarks.part(51).y)
    p57 = (landmarks.part(57).x, landmarks.part(57).y)
    p49 = (landmarks.part(49).x, landmarks.part(49).y)
    p55 = (landmarks.part(55).x, landmarks.part(55).y)
    p53 = (landmarks.part(53).x, landmarks.part(53).y)
    p59 = (landmarks.part(59).x, landmarks.part(59).y)
    return (euclidean_dist(p51, p57) + euclidean_dist(p53, p59) + euclidean_dist(p49, p55)) / (3 * euclidean_dist(p48, p54))

def smile_aspect_ratio(landmarks):
    p48 = (landmarks.part(48).x, landmarks.part(48).y)
    p54 = (landmarks.part(54).x, landmarks.part(54).y)
    p51 = (landmarks.part(51).x, landmarks.part(51).y)
    p57 = (landmarks.part(57).x, landmarks.part(57).y)
    return euclidean_dist(p48, p54) / euclidean_dist(p51, p57)

MOUTH_AR_THRESH = 0.6
SMILE_AR_THRESH = 1.8

# Load predictor dlib
PREDICTOR_PATH = "shape_predictor_68_face_landmarks.dat"
predictor = None
detector_dlib = None
if os.path.exists(PREDICTOR_PATH):
    detector_dlib = dlib.get_frontal_face_detector()
    predictor = dlib.shape_predictor(PREDICTOR_PATH)
haar = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

# -----------------------
# Gesture (dari v4 + "Hi")
# -----------------------
mp_hands = mp.solutions.hands
hands = mp_hands.Hands()
mp_draw = mp.solutions.drawing_utils

def recognize_gesture(hand_landmarks):
    ujung_jempol = hand_landmarks.landmark[mp_hands.HandLandmark.THUMB_TIP]
    ujung_telunjuk = hand_landmarks.landmark[mp_hands.HandLandmark.INDEX_FINGER_TIP]
    ujung_tengah = hand_landmarks.landmark[mp_hands.HandLandmark.MIDDLE_FINGER_TIP]
    ujung_manis = hand_landmarks.landmark[mp_hands.HandLandmark.RING_FINGER_TIP]
    ujung_kelingking = hand_landmarks.landmark[mp_hands.HandLandmark.PINKY_TIP]

    pangkal_jempol = hand_landmarks.landmark[mp_hands.HandLandmark.THUMB_MCP]
    pangkal_telunjuk = hand_landmarks.landmark[mp_hands.HandLandmark.INDEX_FINGER_MCP]
    pangkal_tengah = hand_landmarks.landmark[mp_hands.HandLandmark.MIDDLE_FINGER_MCP]
    pangkal_manis = hand_landmarks.landmark[mp_hands.HandLandmark.RING_FINGER_MCP]
    pangkal_kelingking = hand_landmarks.landmark[mp_hands.HandLandmark.PINKY_MCP]

    if (ujung_jempol.y < ujung_telunjuk.y and ujung_jempol.y < ujung_tengah.y and
        ujung_jempol.y < ujung_manis.y and ujung_jempol.y < ujung_kelingking.y):
        return "Thumbs Up"
    elif (ujung_telunjuk.y < pangkal_telunjuk.y and ujung_tengah.y < pangkal_tengah.y and
          ujung_manis.y > pangkal_manis.y and ujung_kelingking.y > pangkal_kelingking.y):
        return "Peace"
    elif (ujung_telunjuk.y < pangkal_telunjuk.y and ujung_kelingking.y < pangkal_kelingking.y and
          ujung_tengah.y > pangkal_tengah.y and ujung_manis.y > pangkal_manis.y):
        return "Rock"
    elif (abs(ujung_jempol.x - ujung_telunjuk.x) < 0.05 and abs(ujung_jempol.y - ujung_telunjuk.y) < 0.05 and
          ujung_tengah.y < pangkal_tengah.y and ujung_manis.y < pangkal_manis.y and ujung_kelingking.y < pangkal_kelingking.y):
        return "OK"
    elif (ujung_jempol.y < pangkal_jempol.y and ujung_telunjuk.y < pangkal_telunjuk.y and
          ujung_tengah.y < pangkal_tengah.y and ujung_manis.y < pangkal_manis.y and ujung_kelingking.y < pangkal_kelingking.y):
        return "Hi"
    return "Gesture tidak diketahui"

def detect_hand_gesture(frame):
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = hands.process(rgb)
    gesture_found = None
    if results.multi_hand_landmarks:
        for hand_landmarks in results.multi_hand_landmarks:
            gesture_found = recognize_gesture(hand_landmarks)
    return gesture_found

# -----------------------
# LBPH Face Recognizer
# -----------------------
def train_recognizer(dataset_dir=DATASET_DIR):
    if not os.path.exists(dataset_dir): return None, {}
    faces, labels, label_ids, current_id = [], [], {}, 0
    for name in sorted(os.listdir(dataset_dir)):
        path = os.path.join(dataset_dir, name)
        if not os.path.isdir(path): continue
        label_ids[name] = current_id
        for img in os.listdir(path):
            gray = cv2.imread(os.path.join(path, img), cv2.IMREAD_GRAYSCALE)
            if gray is not None:
                faces.append(cv2.resize(gray, IMG_SIZE))
                labels.append(current_id)
        current_id += 1
    if not faces: return None, {}
    rec = cv2.face.LBPHFaceRecognizer_create()
    rec.train(faces, np.array(labels))
    return rec, {v: k for k, v in label_ids.items()}

recognizer, id_to_name = train_recognizer()
last_absen = {}

# -----------------------
# Excel handler
# -----------------------
def append_to_excel(name, waktu, mouth, gesture):
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nama", "Waktu", "Mulut", "Gesture"])
    ws.append([name, waktu, mouth, gesture])
    wb.save(EXCEL_PATH)

# -----------------------
# Kombinasi acak (mulut + gesture)
# -----------------------
mouth_options = ["Senyum", "Mulut terbuka"]
gesture_options = ["Thumbs Up", "Peace", "Rock", "OK", "Hi"]
chosen_mouth = random.choice(mouth_options)
chosen_gesture = random.choice(gesture_options)
print(f"[INFO] Kombinasi: Mulut={chosen_mouth}, Gesture={chosen_gesture}")

# -----------------------
# Fungsi utama deteksi (dipanggil dari web)
# -----------------------
def detect_frame(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    mouth_detected = None
    gesture_found = detect_hand_gesture(frame)

    if detector_dlib and predictor:
        faces = detector_dlib(gray)
        for face in faces:
            landmarks = predictor(gray, face)
            mar = mouth_aspect_ratio(landmarks)
            sar = smile_aspect_ratio(landmarks)
            if sar > SMILE_AR_THRESH: mouth_detected = "Senyum"
            elif mar > MOUTH_AR_THRESH: mouth_detected = "Mulut terbuka"

            x, y, w, h = face.left(), face.top(), face.width(), face.height()
            roi = gray[y:y+h, x:x+w]
            roi = cv2.resize(roi, IMG_SIZE)
            identified_name = "Unknown"
            if recognizer:
                try:
                    label, conf = recognizer.predict(roi)
                    identified_name = id_to_name.get(label, "Unknown")
                except: pass

            now = datetime.now()
            if (mouth_detected == chosen_mouth and gesture_found == chosen_gesture):
                last_time = last_absen.get(identified_name)
                if not last_time or (now - last_time) >= timedelta(minutes=COOLDOWN):
                    append_to_excel(identified_name, now.strftime("%Y-%m-%d %H:%M:%S"), chosen_mouth, chosen_gesture)
                    last_absen[identified_name] = now
                    return gesture_found, mouth_detected, identified_name, "✅ Absensi berhasil"
                else:
                    return gesture_found, mouth_detected, identified_name, "⏳ Tunggu cooldown"
            else:
                return gesture_found, mouth_detected, identified_name, "Belum sesuai instruksi"
    return gesture_found, mouth_detected, "Unknown", "Tidak ada wajah terdeteksi"

# -----------------------
# ROUTES FLASK
# -----------------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/detect', methods=['POST'])
def detect():
    data = request.get_json()
    img_data = base64.b64decode(data['image'].split(',')[1])
    nparr = np.frombuffer(img_data, np.uint8)
    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    gesture, mouth, name, status = detect_frame(frame)
    return jsonify({"gesture": gesture, "mouth": mouth, "name": name, "status": status})

@app.route('/instruction')
def instruction():
    return jsonify({
        "mouth": chosen_mouth,
        "gesture": chosen_gesture
    })



if __name__ == '__main__':
    app.run(debug=True)
